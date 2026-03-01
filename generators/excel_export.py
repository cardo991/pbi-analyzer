"""Excel export generator using openpyxl."""

import io

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def generate_excel(project_name: str, score: dict, findings: list,
                   documentation: dict, kpi_suggestions: list,
                   dax_improvements: list) -> io.BytesIO | None:
    """Generate an Excel workbook with analysis results.

    Returns a BytesIO object ready for send_file(), or None if openpyxl unavailable.
    """
    if not HAS_OPENPYXL:
        return None

    wb = Workbook()

    # Styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="0077CC", end_color="0077CC", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def write_header(ws, row, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    # --- Sheet 1: Summary ---
    ws = wb.active
    ws.title = "Summary"
    overview = documentation.get("overview", {})
    summary_data = [
        ("Project", project_name),
        ("Total Score", score.get("total_score", 0)),
        ("Grade", score.get("grade", "F")),
        ("Total Tables", overview.get("total_tables", 0)),
        ("Total Columns", overview.get("total_columns", 0)),
        ("Total Measures", overview.get("total_measures", 0)),
        ("Total Relationships", overview.get("total_relationships", 0)),
        ("Total Pages", overview.get("total_pages", 0)),
        ("Calculated Columns", overview.get("total_calc_columns", 0)),
    ]
    for cat, cat_score in score.get("category_scores", {}).items():
        summary_data.append((f"Score: {cat}", cat_score))

    write_header(ws, 1, ["Property", "Value"])
    for i, (prop, val) in enumerate(summary_data, 2):
        ws.cell(row=i, column=1, value=prop)
        ws.cell(row=i, column=2, value=val)
    auto_width(ws)

    # --- Sheet 2: Findings ---
    ws = wb.create_sheet("Findings")
    write_header(ws, 1, ["Rule ID", "Category", "Severity", "Message", "Location"])
    findings_data = findings if isinstance(findings, list) else []
    for i, f in enumerate(findings_data, 2):
        if hasattr(f, 'rule_id'):
            ws.cell(row=i, column=1, value=f.rule_id)
            ws.cell(row=i, column=2, value=f.category)
            ws.cell(row=i, column=3, value=f.severity)
            ws.cell(row=i, column=4, value=f.message)
            ws.cell(row=i, column=5, value=f.location)
        elif isinstance(f, dict):
            ws.cell(row=i, column=1, value=f.get("rule_id", ""))
            ws.cell(row=i, column=2, value=f.get("category", ""))
            ws.cell(row=i, column=3, value=f.get("severity", ""))
            ws.cell(row=i, column=4, value=f.get("message", ""))
            ws.cell(row=i, column=5, value=f.get("location", ""))

    # Color code severity
    severity_fills = {
        "error": PatternFill(start_color="FFE2E2", end_color="FFE2E2", fill_type="solid"),
        "warning": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
        "info": PatternFill(start_color="DBE5F1", end_color="DBE5F1", fill_type="solid"),
    }
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
        for cell in row:
            if cell.value in severity_fills:
                for c in ws[cell.row]:
                    c.fill = severity_fills[cell.value]
    auto_width(ws)

    # --- Sheet 3: Tables ---
    ws = wb.create_sheet("Tables")
    write_header(ws, 1, ["Table", "Columns", "Measures", "Mode", "Hidden"])
    for i, t in enumerate(documentation.get("tables", []), 2):
        ws.cell(row=i, column=1, value=t.get("name", ""))
        ws.cell(row=i, column=2, value=t.get("column_count", 0))
        ws.cell(row=i, column=3, value=t.get("measure_count", 0))
        ws.cell(row=i, column=4, value=t.get("mode", ""))
        ws.cell(row=i, column=5, value="Yes" if t.get("is_hidden") else "No")
    auto_width(ws)

    # --- Sheet 4: Measures ---
    ws = wb.create_sheet("Measures")
    write_header(ws, 1, ["Measure", "Table", "Expression", "Format", "Hidden"])
    for i, m in enumerate(documentation.get("measures", []), 2):
        ws.cell(row=i, column=1, value=m.get("name", ""))
        ws.cell(row=i, column=2, value=m.get("table", ""))
        ws.cell(row=i, column=3, value=m.get("expression", ""))
        ws.cell(row=i, column=4, value=m.get("format_string", ""))
        ws.cell(row=i, column=5, value="Yes" if m.get("is_hidden") else "No")
    auto_width(ws)

    # --- Sheet 5: Relationships ---
    ws = wb.create_sheet("Relationships")
    write_header(ws, 1, ["From Table", "From Column", "To Table", "To Column", "Cardinality", "Cross Filter", "Active"])
    for i, r in enumerate(documentation.get("relationships", []), 2):
        ws.cell(row=i, column=1, value=r.get("from_table", ""))
        ws.cell(row=i, column=2, value=r.get("from_column", ""))
        ws.cell(row=i, column=3, value=r.get("to_table", ""))
        ws.cell(row=i, column=4, value=r.get("to_column", ""))
        ws.cell(row=i, column=5, value=f"{r.get('from_cardinality', '')}:{r.get('to_cardinality', '')}")
        ws.cell(row=i, column=6, value=r.get("cross_filtering", ""))
        ws.cell(row=i, column=7, value="Yes" if r.get("is_active") else "No")
    auto_width(ws)

    # --- Sheet 6: Data Sources ---
    ws = wb.create_sheet("Data Sources")
    write_header(ws, 1, ["Table", "Source Type", "Detail", "Mode"])
    for i, s in enumerate(documentation.get("sources", []), 2):
        ws.cell(row=i, column=1, value=s.get("table", ""))
        ws.cell(row=i, column=2, value=s.get("source_type", ""))
        ws.cell(row=i, column=3, value=s.get("detail", ""))
        ws.cell(row=i, column=4, value=s.get("mode", ""))
    auto_width(ws)

    # --- Sheet 7: KPI Suggestions ---
    ws = wb.create_sheet("KPI Suggestions")
    write_header(ws, 1, ["KPI Name", "Description", "DAX Expression", "Format"])
    for i, kpi in enumerate(kpi_suggestions, 2):
        ws.cell(row=i, column=1, value=kpi.get("name", ""))
        ws.cell(row=i, column=2, value=kpi.get("description", ""))
        ws.cell(row=i, column=3, value=kpi.get("dax", ""))
        ws.cell(row=i, column=4, value=kpi.get("format", ""))
    auto_width(ws)

    # Write to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
